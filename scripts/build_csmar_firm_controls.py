#!/usr/bin/env python3
"""Build firm-year controls for the IDC × innovation resilience pilot."""

from __future__ import annotations

import json
import math
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


BASE = Path("/Users/mac/computerscience")
CSMAR = BASE / "第三方资料/01_数据资源/国泰安/第三方数据资源/上市公司财务信息"
OUT = BASE / "23选题探索/T10/outputs"

BALANCE = CSMAR / "资产负债表094840212(仅供四川大学使用)/FS_Combas.xlsx"
AF_ACTUAL = CSMAR / "AF_Actual.xlsx"
CASHFLOW_ZIP = CSMAR / "现金流量表(间接法)105527555(仅供四川大学使用).zip"
CASHFLOW_MEMBER = "FS_Comscfi.xlsx"


def clean_symbol(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .str.extract(r"(\d+)", expand=False)
        .fillna("")
        .str.zfill(6)
        .replace("000000", "")
    )


def to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def year_from_date(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce").dt.year


def parse_year(value) -> tuple[float, str]:
    if isinstance(value, datetime):
        return value.year, value.strftime("%m-%d")
    text = "" if value is None else str(value)
    dt = pd.to_datetime(text, errors="coerce")
    if pd.isna(dt):
        return math.nan, ""
    return int(dt.year), dt.strftime("%m-%d")


def read_xlsx_selected_rows(source, cols: list[str]) -> pd.DataFrame:
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    if ws.max_column == 1:
        ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)
    headers = [str(v) if v is not None else "" for v in next(rows)]
    col_to_idx = {name: headers.index(name) for name in cols}
    # Skip CSMAR Chinese-name and unit rows.
    next(rows, None)
    next(rows, None)
    records = []
    wanted = [(name, col_to_idx[name]) for name in cols]
    for row in rows:
        records.append({name: row[idx] if idx < len(row) else None for name, idx in wanted})
    wb.close()
    return pd.DataFrame.from_records(records, columns=cols)


def read_balance() -> pd.DataFrame:
    cols = ["Stkcd", "Accper", "Typrep", "A001000000", "A002000000", "A003000000"]
    df = read_xlsx_selected_rows(BALANCE, cols)
    df = df[df["Typrep"].astype(str).eq("A")]
    df["symbol"] = clean_symbol(df["Stkcd"])
    parsed = df["Accper"].map(parse_year)
    df["year"] = parsed.map(lambda x: x[0])
    df["mmdd"] = parsed.map(lambda x: x[1])
    df = df[df["mmdd"].eq("12-31")]
    for col in ["A001000000", "A002000000", "A003000000"]:
        df[col] = to_num(df[col])
    df = df[(df["symbol"] != "") & df["year"].notna()]
    df = df.sort_values(["symbol", "year", "Accper"]).drop_duplicates(["symbol", "year"], keep="last")
    out = df[["symbol", "year", "A001000000", "A002000000", "A003000000"]].rename(
        columns={
            "A001000000": "total_assets",
            "A002000000": "total_liabilities",
            "A003000000": "total_equity",
        }
    )
    out["ln_assets"] = np.log(out["total_assets"].where(out["total_assets"] > 0))
    out["lev"] = out["total_liabilities"] / out["total_assets"]
    return out


def read_actual() -> pd.DataFrame:
    cols = ["Stkcd", "Ddate", "Mturnover", "Mnetpro", "BM", "ROA", "TotalAssetsTurnover"]
    df = pd.read_excel(AF_ACTUAL, usecols=cols, dtype=str)
    df = df.iloc[2:].copy()
    df["symbol"] = clean_symbol(df["Stkcd"])
    df["year"] = year_from_date(df["Ddate"])
    df = df[pd.to_datetime(df["Ddate"], errors="coerce").dt.strftime("%m-%d").eq("12-31")]
    for col in ["Mturnover", "Mnetpro", "BM", "ROA", "TotalAssetsTurnover"]:
        df[col] = to_num(df[col])
    df = df[(df["symbol"] != "") & df["year"].notna()]
    df = df.sort_values(["symbol", "year", "Ddate"]).drop_duplicates(["symbol", "year"], keep="last")
    out = df[["symbol", "year", "Mturnover", "Mnetpro", "BM", "ROA", "TotalAssetsTurnover"]].rename(
        columns={
            "Mturnover": "revenue",
            "Mnetpro": "net_profit",
            "BM": "bm",
            "ROA": "roa",
            "TotalAssetsTurnover": "asset_turnover",
        }
    )
    out = out.sort_values(["symbol", "year"])
    lag_rev = out.groupby("symbol")["revenue"].shift(1)
    out["growth"] = (out["revenue"] - lag_rev) / lag_rev.abs()
    return out


def read_cashflow() -> pd.DataFrame:
    cols = ["Stkcd", "Accper", "Typrep", "D000100000"]
    with zipfile.ZipFile(CASHFLOW_ZIP) as zf:
        data = BytesIO(zf.read(CASHFLOW_MEMBER))
    df = read_xlsx_selected_rows(data, cols)
    df = df[df["Typrep"].astype(str).eq("A")]
    df["symbol"] = clean_symbol(df["Stkcd"])
    parsed = df["Accper"].map(parse_year)
    df["year"] = parsed.map(lambda x: x[0])
    df["mmdd"] = parsed.map(lambda x: x[1])
    df = df[df["mmdd"].eq("12-31")]
    df["operating_cf"] = to_num(df["D000100000"])
    df = df[(df["symbol"] != "") & df["year"].notna()]
    df = df.sort_values(["symbol", "year", "Accper"]).drop_duplicates(["symbol", "year"], keep="last")
    return df[["symbol", "year", "operating_cf"]]


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    balance = read_balance()
    actual = read_actual()
    cashflow = read_cashflow()

    controls = balance.merge(actual, on=["symbol", "year"], how="outer")
    controls = controls.merge(cashflow, on=["symbol", "year"], how="outer")
    controls["cfoa"] = controls["operating_cf"] / controls["total_assets"]

    keep = [
        "symbol",
        "year",
        "ln_assets",
        "lev",
        "roa",
        "growth",
        "cfoa",
        "bm",
        "asset_turnover",
        "total_assets",
        "total_liabilities",
        "revenue",
        "operating_cf",
    ]
    controls = controls[keep].sort_values(["symbol", "year"])
    out_csv = OUT / "csmar_firm_controls_annual.csv"
    controls.to_csv(out_csv, index=False, encoding="utf-8-sig")

    report = {
        "output_csv": str(out_csv),
        "rows": int(len(controls)),
        "firms": int(controls["symbol"].nunique()),
        "year_min": int(controls["year"].min()),
        "year_max": int(controls["year"].max()),
        "nonmissing": {col: int(controls[col].notna().sum()) for col in keep[2:9]},
    }
    (OUT / "csmar_firm_controls_build_report_20260523.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
